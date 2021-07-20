#! /usr/bin/python3
# Program that will create spreadsheet which will contain multiplication table
# Usage: ./ch_12_multiplication <number greater than zero>

import sys, openpyxl
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter

if len(sys.argv) != 2:
    print("Usage: ./ch_12_multiplication <number>")
    sys.exit()

wb = openpyxl.Workbook()
sheet = wb["Sheet"]
boldFont = Font(size = 12, bold = True)

# create scale
for i in range(1, int(sys.argv[1]) + 1):
    sheet["A" + str(i+1)] = i
    sheet["A" + str(i+1)].font = boldFont

    sheet.cell(row = 1, column = i + 1).value = i
    sheet[get_column_letter(i + 1) + "1"].font = boldFont

# calculate values
for i in range(1, int(sys.argv[1]) + 1):
    for j in range(1, int(sys.argv[1]) + 1):
        sheet.cell(row = i + 1, column = j + 1).value = i * j

wb.save("multiplication_table.xlsx")
